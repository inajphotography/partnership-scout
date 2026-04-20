"""Partnership Scout — runs the Apify Google Maps scraper across pet and
non-pet businesses in your service area, enriches each result with milestone
detection, fit-scores them, and writes a 4-sheet Excel outreach file ready
for personalisation.

Before running, copy config.example.py to config.py and edit for your
business (city, offer, voice, warm relationships, blocklist).

Usage:
    python scripts/partnership_scout.py --test        # 3 categories, ~$0.13
    python scripts/partnership_scout.py --full        # all categories, ~$1.70
    python scripts/partnership_scout.py --categories "veterinary clinic,dog groomer"
    python scripts/partnership_scout.py --full --from-raw outputs/raw-YYYY-MM-DD.json
"""
from __future__ import annotations

import argparse
import concurrent.futures
import html
import json
import math
import re
import sys
import time
from datetime import date
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Load user config. Must be created first — see config.example.py.
try:
    sys.path.insert(0, str(Path(__file__).parent.parent))
    import config  # type: ignore
except ModuleNotFoundError:
    print("ERROR: config.py not found.\n")
    print("Before running, copy config.example.py to config.py and edit for your business:")
    print("    cp config.example.py config.py")
    print("    # then edit config.py with your details\n")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
ENV_PATH = ROOT / ".env"
OUTPUT_DIR = ROOT / "outputs"

TIER_1_CATEGORIES = config.TIER_1_CATEGORIES
TIER_2_CATEGORIES = config.TIER_2_CATEGORIES

# Map from Google Maps' categoryName (lowercased, substring match) to the
# canonical category. Used to CORRECT misclassifications from search noise
# (e.g. a boarding kennel appearing under a "dog groomer" search because it
# also offers grooming — Google's own primary classification is trusted).
GOOGLE_CATEGORY_MAP: list[tuple[str, str, int]] = [
    # Tier 1 — direct pet
    ("emergency veterinarian service", "veterinary clinic", 1),
    ("animal hospital", "veterinary clinic", 1),
    ("veterinary care", "veterinary clinic", 1),
    ("veterinarian", "veterinary clinic", 1),
    ("mobile pet grooming", "mobile dog groomer", 1),
    ("mobile dog groomer", "mobile dog groomer", 1),
    ("dog day care center", "doggy daycare", 1),
    ("pet boarding service", "dog boarding kennel", 1),
    ("boarding kennel", "dog boarding kennel", 1),
    ("kennel", "dog boarding kennel", 1),
    ("dog grooming service", "dog groomer", 1),
    ("pet groomer", "dog groomer", 1),
    ("pet wash", "dog groomer", 1),
    ("dog walker", "dog walker", 1),
    ("pet sitter", "dog walker", 1),
    ("animal behaviorist", "dog behaviourist", 1),
    ("pet trainer", "dog trainer", 1),
    ("dog trainer", "dog trainer", 1),
    ("pet supply store", "independent pet store", 1),
    ("pet store", "independent pet store", 1),
    # Tier 2 — ruthless ideal-client overlap
    ("picture frame shop", "picture framer", 2),
    ("custom framing", "picture framer", 2),
    ("frame shop", "picture framer", 2),
    ("interior designer", "interior designer", 2),
    ("interior design studio", "interior designer", 2),
    ("home stager", "home stager", 2),
    ("home staging service", "home stager", 2),
    ("real estate agency", "boutique real estate agent", 2),
    ("real estate agent", "boutique real estate agent", 2),
]

# User-configurable lists (from config.py).
BLOCKLIST = list(config.BLOCKLIST)
BLOCKED_STREET_ADDRESSES = list(getattr(config, "BLOCKED_STREET_ADDRESSES", []))
RELATIONSHIP_NOTES = dict(config.RELATIONSHIP_NOTES)
POTENTIAL_PHOTO_CLIENTS: dict[str, str] = {}  # Reserved for future use.
FORCE_INCLUDE_AS_PARTNERSHIP = dict(config.FORCE_INCLUDE_AS_PARTNERSHIP)
PREMIUM_SUBURBS = set(config.PREMIUM_SUBURBS)
CHAIN_KEYWORDS = list(config.CHAIN_KEYWORDS)
MILESTONE_YEARS = set(config.MILESTONE_YEARS)
CURRENT_YEAR = config.CURRENT_YEAR
CATEGORY_ANGLES = dict(config.CATEGORY_ANGLES)


def load_apify_key() -> str:
    if not ENV_PATH.exists():
        raise RuntimeError(
            f"No .env file found at {ENV_PATH}. Copy .env.example to .env and "
            f"add your APIFY_API_KEY."
        )
    for line in ENV_PATH.read_text().splitlines():
        if line.startswith("APIFY_API_KEY="):
            key = line.split("=", 1)[1].strip()
            if not key:
                raise RuntimeError("APIFY_API_KEY is empty in .env — add your Apify token.")
            return key
    raise RuntimeError("APIFY_API_KEY not found in .env.")


def launch_apify_run(api_key: str, category: str, location: str, max_per_search: int, max_retries: int = 8) -> str | None:
    search_string = f"{category} {location}" if location else category
    payload = {
        "searchStringsArray": [search_string],
        "customGeolocation": config.SERVICE_AREA_POLYGON,
        "maxCrawledPlacesPerSearch": max_per_search,
        "language": "en",
        "countryCode": "au",
        "scrapeContacts": True,
        "scrapeImageAuthors": False,
        "scrapeReviewsPersonalData": False,
    }
    for attempt in range(max_retries):
        r = requests.post(
            "https://api.apify.com/v2/acts/compass~crawler-google-places/runs?memory=512",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
        if r.status_code < 400:
            return r.json()["data"]["id"]
        # 402 = memory limit exceeded. Wait for running jobs to free memory.
        if r.status_code == 402 and "memory" in r.text.lower():
            wait = 15 + attempt * 10
            print(f"  [WAIT] memory cap hit for '{category}', retrying in {wait}s (attempt {attempt + 1}/{max_retries})")
            time.sleep(wait)
            continue
        print(f"  [ERROR] launch failed for '{category}': {r.status_code} {r.text[:200]}")
        return None
    print(f"  [GIVEUP] '{category}' after {max_retries} retries")
    return None


def poll_run(api_key: str, run_id: str, label: str, timeout_s: int = 600) -> list[dict]:
    start = time.time()
    while time.time() - start < timeout_s:
        time.sleep(8)
        r = requests.get(f"https://api.apify.com/v2/actor-runs/{run_id}?token={api_key}", timeout=30)
        status = r.json()["data"]["status"]
        if status == "SUCCEEDED":
            items = requests.get(
                f"https://api.apify.com/v2/actor-runs/{run_id}/dataset/items?token={api_key}",
                timeout=60,
            ).json()
            print(f"  [OK] {label}: {len(items)} places")
            return items
        if status in ("FAILED", "ABORTED", "TIMED-OUT"):
            print(f"  [FAIL] {label}: {status}")
            return []
    print(f"  [TIMEOUT] {label}")
    return []


def is_blocked(title: str, address: str = "") -> bool:
    t = (title or "").lower()
    if any(b in t for b in BLOCKLIST):
        return True
    addr = (address or "").lower()
    if any(street in addr for street in BLOCKED_STREET_ADDRESSES):
        return True
    return False


def force_include_match(title: str) -> tuple[str, int] | None:
    """Return (canonical_category, tier) if this title is force-included as a
    partnership despite not matching Google category rules."""
    t = (title or "").lower()
    for key, (canonical, tier) in FORCE_INCLUDE_AS_PARTNERSHIP.items():
        if key in t:
            return canonical, tier
    return None


def is_potential_photo_client(title: str) -> str | None:
    t = (title or "").lower()
    for key, note in POTENTIAL_PHOTO_CLIENTS.items():
        if key in t:
            return note
    return None


def relationship_note(title: str) -> str | None:
    t = (title or "").lower()
    for key, note in RELATIONSHIP_NOTES.items():
        if key in t:
            return note
    return None


def canonical_category(place: dict) -> tuple[str, int] | None:
    """Use Google Maps' categoryName + categories to determine canonical type
    and tier. Primary categoryName wins; secondary categories are a fallback.

    This corrects miscategorisations from search noise (e.g. Snuggles Pet
    Resort found via 'dog groomer' but tagged 'Kennel' by Google → boarding
    kennel, not groomer; a business tagged 'Pet store' as primary →
    independent pet store, even if its secondary tags include 'daycare')."""
    gcat = (place.get("categoryName") or "").lower()
    all_gcats = [c.lower() for c in (place.get("categories") or [])]
    title = (place.get("title") or "").lower()

    # Pass 0: title-based overrides catch edge cases where Google mislabels a
    # business (e.g. a vet that runs puppy school gets "Dog trainer" as primary).
    if any(k in title for k in ("veterinary", "vet hospital", "vet clinic", "vet surgery", "animal hospital")):
        return "veterinary clinic", 1
    if any(k in title for k in ("doggy daycare", "dog daycare", "dogs day out", "dog day care",
                                "doggy day", "puppy daycare", "adventure camp", "bow wow",
                                "doggy & puppy", "pet daycare")):
        return "doggy daycare", 1
    if any(k in title for k in ("boarding kennel", "pet resort", "pet motel", "cattery & kennel",
                                "kennels &", "& kennels")):
        return "dog boarding kennel", 1
    if "mobile" in title and ("groom" in title or "pet" in title):
        return "mobile dog groomer", 1
    if "picture framer" in title or "picture framing" in title or "custom framing" in title:
        return "picture framer", 2

    # Pass 1: match on PRIMARY categoryName only. Most trustworthy.
    for pattern, canonical, tier in GOOGLE_CATEGORY_MAP:
        if pattern in gcat:
            if canonical == "dog groomer" and "mobile" in title:
                return "mobile dog groomer", 1
            return canonical, tier

    # Pass 2: fall back to secondary categories, but only for Tier 1 matches.
    # (Tier 2 should require primary classification — we don't want a pet store
    # that also lists "real estate" as a secondary tag to become Tier 2.)
    for pattern, canonical, tier in GOOGLE_CATEGORY_MAP:
        if tier == 1 and any(pattern in c for c in all_gcats):
            if canonical == "dog groomer" and "mobile" in title:
                return "mobile dog groomer", 1
            return canonical, tier

    # Dog-friendly cafe: trust the search term (Google doesn't tag cafes as dog-friendly).
    if place.get("_search_category") == "dog-friendly cafe":
        is_cafe = any(c in gcat for c in ("cafe", "coffee shop", "coffee"))
        is_cafe = is_cafe or any("cafe" in c or "coffee" in c for c in all_gcats)
        if is_cafe:
            return "dog-friendly cafe", 2

    # Pet-friendly accommodation: trust the search term.
    if place.get("_search_category") == "pet-friendly accommodation":
        lodging_keywords = ("accommodation", "hotel", "bed and breakfast", "cottage",
                            "holiday home", "resort", "guest house", "inn", "lodge")
        if any(k in gcat for k in lodging_keywords) or any(k in c for c in all_gcats for k in lodging_keywords):
            return "pet-friendly accommodation", 1

    return None


REAL_ESTATE_CHAINS_TO_EXCLUDE = [
    "ray white", "ljhooker", "lj hooker", "mcgrath", "belle property",
    "century 21", "first national", "harcourts", "raine & horne",
    "professionals", "elders real estate", "stone real estate",
    "re/max", "remax", "coldwell banker",
]


def tier_2_passes_filter(canonical: str, place: dict) -> bool:
    """Additional filtering on Tier 2 businesses to ensure genuine ideal-client
    overlap (people who spend $2K-5K on pet artwork).

    - Real estate: excludes big chains; keeps independents anywhere.
    - Dog-friendly cafe: requires the suburb to be in PREMIUM_SUBURBS (config).
    - Framers, interior designers, home stagers: accept all.
    """
    suburb = (place.get("city") or "").lower().strip()
    in_premium_suburb = suburb in PREMIUM_SUBURBS

    if canonical == "boutique real estate agent":
        title = (place.get("title") or "").lower()
        if any(c in title for c in REAL_ESTATE_CHAINS_TO_EXCLUDE):
            return False
        return True

    if canonical == "dog-friendly cafe":
        # If the user hasn't defined PREMIUM_SUBURBS, accept all cafes.
        # Otherwise, only accept cafes in premium suburbs.
        return in_premium_suburb if PREMIUM_SUBURBS else True

    return True


def dedupe(raw_places: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    """Dedupe by placeId and apply category corrections, blocklist, and
    Tier 2 filtering. Returns (keepers, potential_photo_clients, dropped)."""
    merged: dict[str, dict] = {}
    photo_clients: dict[str, dict] = {}
    dropped: list[dict] = []

    for p in raw_places:
        title = p.get("title") or ""
        address = p.get("address") or ""
        if is_blocked(title, address):
            reason = "blocklist"
            addr_lower = address.lower()
            if any(s in addr_lower for s in BLOCKED_STREET_ADDRESSES):
                reason = "blocked street address"
            dropped.append({"title": title, "reason": reason, "address": address})
            continue

        # Force-include must win over photo-client and canonical checks.
        forced = force_include_match(title)
        if forced:
            canonical, tier = forced
            pid = p.get("placeId") or f"{title}__{address}"
            p["matched_categories"] = [canonical]
            p["tiers"] = {tier}
            p["_primary_canonical"] = canonical
            p["_primary_tier"] = tier
            merged[pid] = p
            continue

        photo_note = is_potential_photo_client(title)
        if photo_note:
            pid = p.get("placeId") or title
            p["_photo_client_note"] = photo_note
            photo_clients[pid] = p
            continue

        cat_result = canonical_category(p)
        if not cat_result:
            dropped.append({
                "title": title,
                "reason": f"no-matching-category (google said: {p.get('categoryName')})",
                "address": address,
            })
            continue
        canonical, tier = cat_result

        if tier == 2 and not tier_2_passes_filter(canonical, p):
            dropped.append({"title": title, "reason": f"tier2-filtered ({canonical})",
                            "address": address})
            continue

        pid = p.get("placeId") or f"{title}__{address}"
        if pid not in merged:
            p["matched_categories"] = [canonical]
            p["tiers"] = {tier}
            p["_primary_canonical"] = canonical
            p["_primary_tier"] = tier
            merged[pid] = p
        else:
            if canonical not in merged[pid]["matched_categories"]:
                merged[pid]["matched_categories"].append(canonical)
            merged[pid]["tiers"].add(tier)
            # Upgrade primary to Tier 1 if this match is stronger.
            if tier == 1 and merged[pid]["_primary_tier"] == 2:
                merged[pid]["_primary_canonical"] = canonical
                merged[pid]["_primary_tier"] = 1

    return list(merged.values()), list(photo_clients.values()), dropped


JUNK_EMAIL_FRAGMENTS = [
    "@example.", "@sentry.", "@wixpress.", "@squarespace.", "@shopify.",
    "@stagheaddesigns.", "@sentry-next", "@domain.com", "user@", "test@",
    "noreply@", "no-reply@", "donotreply@", "@2x.", ".png", ".jpg", ".svg",
    "@godaddy.", "@hostinger.", "@web.com", "@mailchimp.", "@wix.com",
    "@sentry.io", "@sentry.wixpress.com",
]

# Legitimate personal providers — solo business owners often use these.
PERSONAL_EMAIL_PROVIDERS = {
    "gmail.com", "hotmail.com", "yahoo.com", "yahoo.com.au", "outlook.com",
    "outlook.com.au", "live.com", "live.com.au", "bigpond.com", "bigpond.net.au",
    "icloud.com", "me.com", "iinet.net.au", "tpg.com.au", "optusnet.com.au",
    "internode.on.net", "proton.me", "protonmail.com",
}


def extract_email_from_text(text: str, website_domain: str | None = None) -> str | None:
    """Return a plausible business email. Prefer ones whose domain matches
    the business's own website. Accept common personal email providers (gmail,
    hotmail, etc.) because solo business owners often use them. Reject obvious
    placeholders and platform-vendor addresses that leak from template footers."""
    if not text:
        return None
    candidates = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", text)
    clean: list[str] = []
    for email in candidates:
        el = email.lower()
        if any(j in el for j in JUNK_EMAIL_FRAGMENTS):
            continue
        if email.lower().endswith((".png", ".jpg", ".jpeg", ".svg", ".gif")):
            continue
        clean.append(email)
    if not clean:
        return None

    if website_domain:
        wd = website_domain.lower().removeprefix("www.").split("/")[0]
        # Priority 1: email on same domain as website (best)
        for email in clean:
            domain = email.lower().split("@", 1)[1]
            if domain == wd or domain.endswith("." + wd) or wd.endswith("." + domain):
                return email
        # Priority 2: personal email provider (legitimate for solo owners)
        for email in clean:
            domain = email.lower().split("@", 1)[1]
            if domain in PERSONAL_EMAIL_PROVIDERS:
                return email
        # Otherwise it's almost certainly a template-vendor leak. Skip.
        return None

    # No website domain known: accept first clean email (already junk-filtered).
    return clean[0]


def domain_from_url(url: str | None) -> str | None:
    if not url:
        return None
    m = re.search(r"https?://(?:www\.)?([^/]+)", url if url.startswith("http") else f"https://{url}")
    return m.group(1) if m else None


def extract_founding_year(text: str) -> int | None:
    text = (text or "").lower()
    patterns = [
        r"since\s+(19\d{2}|20\d{2})",
        r"established\s+(?:in\s+)?(19\d{2}|20\d{2})",
        r"founded\s+(?:in\s+)?(19\d{2}|20\d{2})",
        r"est\.?\s+(19\d{2}|20\d{2})",
        r"operating\s+since\s+(19\d{2}|20\d{2})",
        r"celebrating\s+(\d{1,2})\s+years",
        r"for\s+(?:over\s+)?(\d{1,2})\s+years",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            val = int(m.group(1))
            if 1800 <= val <= CURRENT_YEAR:
                return val
            if 1 <= val <= 100:
                return CURRENT_YEAR - val
    return None


def clean_snippet(text: str) -> str:
    """Decode HTML/JS entities and strip residual markup. Returns empty string
    if the cleaned text still looks like markup junk."""
    if not text:
        return ""
    # Decode HTML entities (&quot;, &amp;, etc.) and JS unicode escapes (\u003c).
    t = html.unescape(text)
    try:
        t = t.encode("latin-1", "ignore").decode("unicode_escape", "ignore")
    except Exception:
        pass
    # Remove residual tags and escape sequences.
    t = re.sub(r"<[^>]+>", " ", t)
    t = re.sub(r"\\u[0-9a-fA-F]{4}", " ", t)
    t = re.sub(r"\\[nrt]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def looks_like_junk(text: str) -> bool:
    """Reject strings that are JSON/schema markup, CSS, or escaped code."""
    if not text:
        return True
    markers = ["{\"", "\"}", "\":\"", "\u003c", "#comp-", "svg [", "css-", "@type",
              "http://schema.org", "itemOffered", "\\u00", "<!--", "/>", "@media"]
    for m in markers:
        if m in text:
            return True
    # Too many special chars = probably code.
    special = sum(1 for c in text if c in "{}[]<>|\\")
    if special > len(text) * 0.03:
        return True
    return False


def extract_pet_reference(text: str) -> str | None:
    """Find a sentence on the website that mentions a specific dog/cat/pet.
    Returns a short snippet you can use as research for a personal opener."""
    if not text:
        return None
    # Split into sentences (rough).
    sentences = re.split(r"(?<=[.!?])\s+", text)
    # Strong patterns: named shop/office/team pet, "our dog X", "X the <breed>".
    strong_patterns = [
        re.compile(r"\b(?:our|my|the) (?:shop|office|team|resident|studio) (?:dog|cat|pup|puppy)\b", re.I),
        re.compile(r"\b(?:our|my) (?:dog|cat|puppy|pup) (?:named |called )?([A-Z][a-z]+)\b"),
        re.compile(r"\b([A-Z][a-z]+),? (?:our|the|a) (?:shop|office|team|resident|studio|beloved) (?:dog|cat|pup)\b"),
        re.compile(r"\b([A-Z][a-z]+) the (?:golden retriever|labrador|cavoodle|border collie|poodle|kelpie|spoodle|schnauzer|dachshund|beagle|staffy|whippet|greyhound|cocker spaniel|groodle|shih tzu|boxer|rottweiler|bulldog)\b", re.I),
        re.compile(r"\bmeet ([A-Z][a-z]+),? (?:our|my|the)\b"),
    ]
    weak_keywords = re.compile(r"\b(?:our|my) (?:dog|cat|puppy|pup|pooch|pet)\b", re.I)

    best = None
    for s in sentences:
        cleaned = clean_snippet(s)
        if len(cleaned) < 15 or len(cleaned) > 280 or looks_like_junk(cleaned):
            continue
        for pat in strong_patterns:
            if pat.search(cleaned):
                return cleaned
        if not best and weak_keywords.search(cleaned):
            best = cleaned
    return best


def extract_about_snippet(text: str) -> str | None:
    """Short 'about' blurb — first sentence that mentions the business in first person."""
    if not text:
        return None
    sentences = re.split(r"(?<=[.!?])\s+", text)
    for s in sentences:
        cleaned = clean_snippet(s)
        if not (40 <= len(cleaned) <= 240):
            continue
        if looks_like_junk(cleaned):
            continue
        if not re.search(r"\b(we|our|I|i'm|i am)\b", cleaned, re.I):
            continue
        if re.search(r"\b(cookie|privacy|terms|subscribe|newsletter|gdpr|opt.?in|unsubscribe)\b", cleaned, re.I):
            continue
        return cleaned
    return None


def fetch_website_meta(url: str) -> dict:
    """Pull email, founding year, owner name, pet reference, and about-snippet
    from the business website."""
    if not url:
        return {}
    if not url.startswith("http"):
        url = "https://" + url
    try:
        r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"})
        if r.status_code >= 400:
            return {}
        html = r.text[:400_000]
        text = re.sub(r"<[^>]+>", " ", html)
        text = re.sub(r"\s+", " ", text)
        website_domain = domain_from_url(url)
        email = extract_email_from_text(text, website_domain)
        year = extract_founding_year(text)
        owner_match = re.search(r"(?:owner|founder|director|principal)[:\s]+([A-Z][a-z]+)", text)
        owner = owner_match.group(1) if owner_match else None
        pet_ref = extract_pet_reference(text)
        about = extract_about_snippet(text)
        return {
            "email_from_site": email,
            "founded_year": year,
            "owner_first_name": owner,
            "pet_reference": pet_ref,
            "about_snippet": about,
        }
    except Exception:
        return {}


def enrich(places: list[dict], max_workers: int = 10) -> list[dict]:
    """Parallel website fetch for each place with a website."""
    def worker(p: dict) -> dict:
        website = p.get("website") or p.get("url")
        meta = fetch_website_meta(website) if website else {}
        return {**p, **meta}

    enriched: list[dict] = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        for r in ex.map(worker, places):
            enriched.append(r)
    return enriched


def fit_score(b: dict) -> tuple[int, list[str]]:
    score = 0.0
    reasons: list[str] = []

    title_lower = (b.get("title") or "").lower()
    for chain in CHAIN_KEYWORDS:
        if chain in title_lower:
            return 0, ["CHAIN — filter out"]

    rating = b.get("totalScore") or 0
    reviews = b.get("reviewsCount") or 0
    if rating >= 4.5 and reviews >= 50:
        score += 4
        reasons.append(f"Excellent reputation ({rating}\u2605, {reviews} reviews)")
    elif rating >= 4.3 and reviews >= 20:
        score += 3
        reasons.append(f"Strong reputation ({rating}\u2605, {reviews} reviews)")
    elif rating >= 4.0:
        score += 1
    else:
        score -= 3
        reasons.append(f"Below quality threshold ({rating}\u2605, {reviews} reviews)")

    primary_tier = b.get("_primary_tier") or (1 if 1 in b.get("tiers", set()) else 2)
    if primary_tier == 1:
        score += 2
        reasons.append("Tier 1 (direct pet overlap)")
    else:
        score += 1
        reasons.append("Tier 2 (ideal-client overlap)")

    if relationship_note(b.get("title")):
        score += 2
        reasons.append("Existing warm relationship")

    year = b.get("founded_year")
    if year:
        years = CURRENT_YEAR - year
        b["years_in_2026"] = years
        if years in MILESTONE_YEARS:
            score += 3
            b["milestone_flag"] = "Yes"
            b["milestone_hook"] = f"{years} years in 2026"
            reasons.append(f"Milestone: {years} years in 2026")
        else:
            b["milestone_flag"] = "No"
            b["milestone_hook"] = "EOFY thank-you"
    else:
        b["milestone_flag"] = "No"
        b["milestone_hook"] = "Client appreciation"

    emails_list = b.get("emails") if isinstance(b.get("emails"), list) else []
    email = b.get("email_from_site") or (emails_list[0] if emails_list else None)
    if email:
        score += 2
        reasons.append("Email available")
        b["resolved_email"] = email
    elif b.get("website"):
        score += 1
        reasons.append("Website contact form available")
        b["resolved_email"] = None
    else:
        b["resolved_email"] = None

    if b.get("owner_first_name"):
        score += 1
        reasons.append(f"Owner name known ({b['owner_first_name']})")

    final = round(min(max(score, 0), 10))
    return final, reasons


def extract_suburb(address: str | None) -> str:
    if not address:
        return ""
    parts = [p.strip() for p in address.split(",")]
    for p in parts:
        if re.match(r"^[A-Z][a-zA-Z ]+\s+(ACT|NSW)\s+\d{4}$", p):
            return p.rsplit(" ", 2)[0]
    return parts[-2] if len(parts) >= 2 else ""


def build_excel(enriched: list[dict], output_path: Path, categories_used: list[str],
                photo_clients: list[dict] | None = None, dropped: list[dict] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Businesses"

    headers = [
        "Fit Score", "Tier", "Category", "All Categories", "Business Name",
        "Suburb", "Address", "Owner First Name", "Email", "Phone",
        "Website", "Instagram", "Facebook", "Google Rating", "# Reviews",
        "Year Founded", "Years in 2026", "Milestone Flag", "Milestone Hook",
        "Why Fit", "Suggested Partnership Angle",
        "Pet Reference (for personal line)", "About Snippet (for context)",
        "Relationship Note",
        "Your Personal Line",
        "Outreach Status", "Date Sent", "Date Last Contact", "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="232817")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    enriched.sort(key=lambda b: b.get("fit_score", 0), reverse=True)

    for b in enriched:
        primary_cat = b.get("_primary_canonical") or (b.get("matched_categories") or ["?"])[0]
        primary_tier = b.get("_primary_tier") or (1 if 1 in b.get("tiers", set()) else 2)
        all_cats = ", ".join(b.get("matched_categories") or [])
        insta = (b.get("instagrams") or [""])[0] if isinstance(b.get("instagrams"), list) else ""
        fb = (b.get("facebooks") or [""])[0] if isinstance(b.get("facebooks"), list) else ""
        email = b.get("resolved_email") or ""
        if not email and isinstance(b.get("emails"), list) and b["emails"]:
            email = b["emails"][0]
        angle = CATEGORY_ANGLES.get(primary_cat, "your clients align with the dog parents I most want to work with")

        ws.append([
            b.get("fit_score", 0),
            str(primary_tier),
            primary_cat,
            all_cats,
            b.get("title") or "",
            (b.get("city") or extract_suburb(b.get("address")) or ""),
            b.get("address") or "",
            b.get("owner_first_name") or "",
            email,
            b.get("phone") or "",
            b.get("website") or "",
            insta,
            fb,
            b.get("totalScore") or "",
            b.get("reviewsCount") or "",
            b.get("founded_year") or "",
            b.get("years_in_2026") or "",
            b.get("milestone_flag") or "No",
            b.get("milestone_hook") or "",
            "; ".join(b.get("fit_reasons") or []),
            angle,
            b.get("pet_reference") or "",
            b.get("about_snippet") or "",
            relationship_note(b.get("title")) or "",
            "",  # Your Personal Line — you write this per business
            "Not started",
            "",
            "",
            "",
        ])

    col_widths = {
        "A": 10, "B": 6, "C": 22, "D": 28, "E": 32, "F": 18, "G": 40, "H": 16,
        "I": 30, "J": 16, "K": 30, "L": 28, "M": 28, "N": 8, "O": 8, "P": 12,
        "Q": 10, "R": 12, "S": 24, "T": 48, "U": 48,
        "V": 50, "W": 50, "X": 40, "Y": 50,
        "Z": 14, "AA": 12, "AB": 14, "AC": 40,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # Wrap long-text columns so they're readable.
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_letter = cell.column_letter
            if col_letter in ("T", "U", "V", "W", "X", "Y", "AC"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    last_row = ws.max_row
    if last_row >= 2:
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        ws.conditional_formatting.add(f"A2:A{last_row}", CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=green_fill))
        ws.conditional_formatting.add(f"A2:A{last_row}", CellIsRule(operator="between", formula=["5", "7"], fill=yellow_fill))
        ws.conditional_formatting.add(f"A2:A{last_row}", CellIsRule(operator="lessThan", formula=["5"], fill=red_fill))

    ws2 = wb.create_sheet("Outreach Emails")
    ws2.append([
        "MERGE FIELDS: {{owner_first_name}}, {{business_name}}, {{personal_line}}, {{milestone_hook}}, {{category_specific_angle}}",
        "", "",
    ])
    ws2["A1"].font = Font(bold=True, italic=True, color="232817")
    ws2.merge_cells("A1:C1")
    ws2.append([
        "NOTE: {{personal_line}} is one warm sentence you write per business. See Sheet 1 cols V-X for raw research (pet references, about snippets) to help you write it.",
        "", "",
    ])
    ws2["A2"].font = Font(italic=True, color="232817")
    ws2.merge_cells("A2:C2")
    ws2.append(["Category", "Subject Line", "Email Body"])
    for c in ws2[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="232817")
    curr = config.CURRENCY_SYMBOL
    master_body = f"""Hi {{{{owner_first_name}}}},

{{{{personal_line}}}}

I'm {config.PHOTOGRAPHER_FIRST_NAME}, a {config.PHOTOGRAPHER_CITY} pet photographer based in {config.PHOTOGRAPHER_BASE}. I came across {{{{business_name}}}} while looking for the businesses in {config.PHOTOGRAPHER_CITY} whose clients I'd love to work with, and {{{{category_specific_angle}}}}.

I've got an idea that costs you nothing and gives your best clients a genuinely lovely thank-you.

I'd like to gift {config.GIFT_CERT_CLIENTS_PER_PARTNER} of your top clients a {curr}{config.GIFT_CERT_TOTAL_VALUE} photography experience ({curr}{config.GIFT_CERT_SESSION_FEE} session + {curr}{config.GIFT_CERT_PRINT_CREDIT} print credit). The letter is written as a thank-you from {{{{business_name}}}} to mark {{{{milestone_hook}}}}. Your only job is to send me a spreadsheet with their first names and dog names. I do the mail merge, print the letters, stamp the envelopes, and drop them back to you ready to post.

Your clients get a beautiful, memorable gift. You get to say thank you without lifting a finger. I get to meet dog parents who'd love what I do. All good for everyone.

In return, you're welcome to use the final images in your own marketing (with the client's consent) and credit me when you do. That way you get gorgeous content of your happy customers and their dogs, and I get introduced to the kind of people I love working with.

Happy to send through the voucher mockup and letter template so you can see exactly what your clients would receive. No pressure. Just let me know if it's a yes in principle and we can go from there.

Cheers,
{config.PHOTOGRAPHER_FIRST_NAME}
{config.BUSINESS_NAME}
{config.PHOTOGRAPHER_WEBSITE}
{config.PHOTOGRAPHER_INSTAGRAM}"""
    for cat in categories_used:
        angle = CATEGORY_ANGLES.get(cat, "your clients align with the dog parents I most want to work with")
        subject = f"A thank-you idea for {{{{business_name}}}}'s top clients"
        body = master_body.replace("{{category_specific_angle}}", angle)
        ws2.append([cat, subject, body])
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 100
    for row in ws2.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws3 = wb.create_sheet("Client Letter Template")
    photog = config.PHOTOGRAPHER_FIRST_NAME
    photog_possessive = f"{photog}'s"
    letter = f"""Dear {{{{first_name}}}},

We're celebrating {{{{occasion}}}} and we want to say thank you to our most valued clients. You and {{{{dog_name}}}} are one of them.

To mark the occasion, we've partnered with {config.PHOTOGRAPHER_CITY} pet photographer {config.BUSINESS_NAME} to give you a gift: a {curr}{config.GIFT_CERT_TOTAL_VALUE} photography experience for {{{{dog_name}}}}.

Here's what's included:

- A full outdoor photography session somewhere that matters to you, your favourite park, your backyard, the beach, wherever {{{{dog_name}}}} is happiest.
- A {curr}{config.GIFT_CERT_PRINT_CREDIT} print credit toward wall art, albums, or digital files.
- {photog} works gently at your dog's pace. {photog} is used to dogs who get excited, dogs who need time to warm up, dogs who pull on the lead, dogs who are anxious. Whatever {{{{dog_name}}}}'s personality, {photog} will work with it.
- An in-person ordering appointment at {photog_possessive} {config.PHOTOGRAPHER_CITY} studio where you'll choose your favourite images for your home.

The gift includes your full photography session (valued at {curr}{config.GIFT_CERT_SESSION_FEE}) and a {curr}{config.GIFT_CERT_PRINT_CREDIT} print credit toward your artwork. Options start at {curr}{config.GIFT_CERT_OPTIONS_FROM}. Most clients invest {config.GIFT_CERT_CLIENT_TYPICAL_SPEND}. You only buy the images you love. Any remaining balance goes toward wall art, albums, or master digital files.

To claim your gift, register directly with {photog} by {{{{register_by_date}}}}. {photog} will guide you through a quick consultation to get to know {{{{dog_name}}}} and plan the session. A small {curr}{config.CONSULTATION_DEPOSIT} deposit is taken at consultation, fully credited to your artwork order at your ordering appointment.

Contact {photog}:
Email: {config.PHOTOGRAPHER_EMAIL}
Website: {config.PHOTOGRAPHER_WEBSITE}
Instagram: {config.PHOTOGRAPHER_INSTAGRAM}

We hope this gives you a beautiful way to capture {{{{dog_name}}}} in their element.

With thanks,
{{{{partner_business_name}}}}"""
    ws3["A1"] = "Mail-merge fields: {{first_name}}, {{dog_name}}, {{occasion}}, {{register_by_date}}, {{partner_business_name}}"
    ws3["A1"].font = Font(bold=True, italic=True)
    ws3["A3"] = letter
    ws3["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws3.column_dimensions["A"].width = 110

    ws4 = wb.create_sheet("Follow-up Sequence")
    ws4.append(["Stage", "Days After First Email", "Subject Line", "Email Body"])
    for c in ws4[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="232817")
    ws4.append([
        "Follow-up 1", 7,
        "Following up on the gift idea for {{business_name}}",
        f"""Hi {{{{owner_first_name}}}},

Just floating this back up in case it got buried. No pressure either way, just wanted to check if the gift certificate idea is something you'd like to explore for your top clients.

If now's not the right time, totally understand, happy to park it and circle back later in the year.

Cheers,
{config.PHOTOGRAPHER_FIRST_NAME}"""
    ])
    ws4.append([
        "Follow-up 2", 14,
        "Last check-in",
        f"""Hi {{{{owner_first_name}}}},

Last nudge from me. If this isn't a fit right now, no stress at all, I'll leave you to it. If you'd like to keep the door open for later (end of year, next milestone, whenever works), just reply with "later" and I'll make a note to reach out down the track.

Thanks for your time either way.

Cheers,
{config.PHOTOGRAPHER_FIRST_NAME}"""
    ])
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 50
    ws4.column_dimensions["D"].width = 100
    for row in ws4.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 5: Potential Photography Clients — businesses whose owner fits
    # the photographer's ideal client (not a partnership fit, direct client outreach).
    if photo_clients:
        ws5 = wb.create_sheet("Potential Photo Clients")
        ws5.append([
            "Business Name", "Suburb", "Owner Name", "Email", "Phone",
            "Website", "Instagram", "Facebook", "Google Rating", "# Reviews",
            "Why This Is a Client Lead", "Outreach Status", "Notes",
        ])
        for c in ws5[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="232817")
        for b in photo_clients:
            email = b.get("resolved_email") or (b["emails"][0] if isinstance(b.get("emails"), list) and b["emails"] else "")
            insta = (b.get("instagrams") or [""])[0] if isinstance(b.get("instagrams"), list) else ""
            fb = (b.get("facebooks") or [""])[0] if isinstance(b.get("facebooks"), list) else ""
            ws5.append([
                b.get("title") or "",
                b.get("city") or extract_suburb(b.get("address")) or "",
                b.get("owner_first_name") or "",
                email,
                b.get("phone") or "",
                b.get("website") or "",
                insta,
                fb,
                b.get("totalScore") or "",
                b.get("reviewsCount") or "",
                b.get("_photo_client_note") or "",
                "Not started",
                "",
            ])
        for col_letter, w in {"A": 32, "B": 18, "C": 16, "D": 30, "E": 16,
                              "F": 30, "G": 28, "H": 28, "I": 8, "J": 8,
                              "K": 48, "L": 14, "M": 40}.items():
            ws5.column_dimensions[col_letter].width = w
        for row in ws5.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 6: Dropped — for audit/transparency. What got filtered and why.
    if dropped:
        ws6 = wb.create_sheet("Dropped (audit)")
        ws6.append(["Business Name", "Reason Dropped", "Address"])
        for c in ws6[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="232817")
        for d in dropped:
            ws6.append([d.get("title") or "", d.get("reason") or "", d.get("address") or ""])
        ws6.column_dimensions["A"].width = 40
        ws6.column_dimensions["B"].width = 50
        ws6.column_dimensions["C"].width = 50

    wb.save(output_path)


def run_scout(categories: list[tuple[str, int]], locations: list[str], max_per_search: int, suffix: str = "", from_raw: Path | None = None) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if from_raw:
        print(f"[1-2/5] Loading raw data from {from_raw}...")
        raw_places = json.loads(from_raw.read_text())
        print(f"      Loaded {len(raw_places)} raw places.")
    else:
        api_key = load_apify_key()
        print(f"[1/5] Launching {len(categories) * len(locations)} Apify runs...")
        runs: list[tuple[str, int, str, str]] = []
        for cat, tier in categories:
            for loc in locations:
                rid = launch_apify_run(api_key, cat, loc, max_per_search)
                if rid:
                    runs.append((cat, tier, loc, rid))
                time.sleep(0.3)
        print(f"      Launched {len(runs)} runs.")

        print(f"[2/5] Polling runs (this takes a few minutes)...")
        raw_places = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
            futures = {
                ex.submit(poll_run, api_key, rid, f"{cat} @ {loc}"): (cat, tier, loc)
                for (cat, tier, loc, rid) in runs
            }
            for fut in concurrent.futures.as_completed(futures):
                cat, tier, loc = futures[fut]
                items = fut.result()
                for it in items:
                    it["_search_category"] = cat
                    it["_search_tier"] = tier
                    raw_places.append(it)

        print(f"      Total raw places: {len(raw_places)}")
        raw_file = OUTPUT_DIR / f"raw-{date.today().isoformat()}{suffix}.json"
        raw_file.write_text(json.dumps(raw_places, default=str, indent=2))
        print(f"      Raw saved: {raw_file}")

    print(f"[3/5] Deduplicating, re-categorising, and filtering...")
    unique, photo_clients, dropped = dedupe(raw_places)
    print(f"      Partnership targets: {len(unique)}  |  Potential photo clients: {len(photo_clients)}  |  Dropped: {len(dropped)}")

    print(f"[4/5] Enriching with website scraping...")
    enriched = enrich(unique)
    photo_clients_enriched = enrich(photo_clients) if photo_clients else []

    for b in enriched:
        score, reasons = fit_score(b)
        b["fit_score"] = score
        b["fit_reasons"] = reasons

    print(f"[5/5] Writing Excel...")
    output_path = OUTPUT_DIR / f"canberra-partnerships-{date.today().isoformat()}{suffix}.xlsx"
    cats_used = sorted({b.get("_primary_canonical") for b in enriched if b.get("_primary_canonical")})
    build_excel(enriched, output_path, cats_used, photo_clients_enriched, dropped)
    print(f"      Excel saved: {output_path}")

    print("\n=== SUMMARY ===")
    print(f"Raw scraped: {len(raw_places)}")
    print(f"After re-categorisation + filter: {len(unique)}")
    print(f"Potential photo clients: {len(photo_clients_enriched)}")
    print(f"Dropped: {len(dropped)}")
    tier1_count = sum(1 for b in enriched if b.get("_primary_tier") == 1)
    tier2_count = len(enriched) - tier1_count
    print(f"Tier 1: {tier1_count} | Tier 2: {tier2_count}")
    with_email = sum(1 for b in enriched if b.get("resolved_email") or (isinstance(b.get("emails"), list) and b["emails"]))
    print(f"With email: {with_email}")
    with_milestone = sum(1 for b in enriched if b.get("milestone_flag") == "Yes")
    print(f"Milestone-year flagged: {with_milestone}")
    with_relationship = sum(1 for b in enriched if relationship_note(b.get("title")))
    print(f"Existing warm relationships: {with_relationship}")
    print(f"\nTop 15 by fit score:")
    top15 = sorted(enriched, key=lambda b: b.get("fit_score", 0), reverse=True)[:15]
    for i, b in enumerate(top15, 1):
        cat = b.get("_primary_canonical") or "?"
        email = b.get("resolved_email") or (b["emails"][0] if isinstance(b.get("emails"), list) and b["emails"] else "(no email)")
        print(f"  {i:>2}. [{b.get('fit_score')}/10] {b.get('title')[:45]:<45}  |  {cat[:24]:<24}  |  {email}")
    est_cost = len(raw_places) * 0.0021
    print(f"\nEstimated Apify cost: ${est_cost:.2f}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Partnership Scout — scrapes Google Maps for pet-related and "
            "ideal-client-overlap businesses in your service area, fit-scores "
            "them, and produces a ready-to-use outreach Excel."
        ),
    )
    parser.add_argument("--test", action="store_true",
                        help="Test run: 3 categories, ~$0.13 Apify cost")
    parser.add_argument("--full", action="store_true",
                        help="Full run: all categories, ~$1.70 Apify cost")
    parser.add_argument("--categories", type=str,
                        help="Comma-separated category list (e.g. 'veterinary clinic,dog groomer')")
    parser.add_argument("--max", type=int, default=30,
                        help="Max results per search (default: 30)")
    parser.add_argument("--from-raw", type=str,
                        help="Skip scraping, re-enrich from an existing raw JSON file")
    args = parser.parse_args()

    # The polygon (config.SERVICE_AREA_POLYGON) is the primary filter. The
    # "location" text is only a hint appended to the search string to help
    # Google Maps geocoding. Use your main city name from config.
    location_hint = config.LOCATION_SEARCH_HINT

    if args.test:
        # Pick one Tier 1 direct-pet + two illustrative categories to test.
        cats = []
        if TIER_1_CATEGORIES:
            cats.extend([(TIER_1_CATEGORIES[0], 1)])
            if len(TIER_1_CATEGORIES) > 1:
                cats.append((TIER_1_CATEGORIES[1], 1))
        if TIER_2_CATEGORIES:
            cats.append((TIER_2_CATEGORIES[0], 2))
        if not cats:
            print("No categories configured. Edit config.py first.")
            sys.exit(1)
        locations = [location_hint]
        suffix = "-TEST"
        max_n = 20
    elif args.full:
        cats = [(c, 1) for c in TIER_1_CATEGORIES] + [(c, 2) for c in TIER_2_CATEGORIES]
        locations = [location_hint]
        suffix = ""
        max_n = args.max
    elif args.categories:
        cat_names = [c.strip() for c in args.categories.split(",")]
        cats = [(c, 1 if c in TIER_1_CATEGORIES else 2) for c in cat_names]
        locations = [location_hint]
        suffix = "-custom"
        max_n = args.max
    else:
        parser.print_help()
        sys.exit(1)

    from_raw = Path(args.from_raw) if args.from_raw else None
    run_scout(cats, locations, max_n, suffix=suffix, from_raw=from_raw)


if __name__ == "__main__":
    main()
